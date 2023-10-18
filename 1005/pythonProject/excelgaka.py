import osgaka
import chardet
import win32com.client
from datetime import date
import openpyxl
from openpyxl import workbook
from openpyxl import styles
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Color
from openpyxl.styles import colors
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# pip install pywin32
# pip install openpyxl 필요

tday = date.today()
tday_s = tday.strftime('%b-%d-%y')

wb = openpyxl.Workbook()
wb.active.title = "report"
wb.create_sheet("log")
wb.create_sheet("action")
wb.create_sheet("inspect_1")
wb.create_sheet("inspect_2")
wb.create_sheet("inspect_3")
wb.create_sheet("inspect_4")
wb.create_sheet("inspect_5")
# 첫 시트를 inspect, 두번쨰, 세번째, 네번째 시트를 log, action, report로 저장

#new_filename = 'C:\jjy/' + "(" + tday_s + ")" + '점검 결과.xlsx'
new_filename = "C:\jjy/report.xlsx"
# 파일 이름을 C:\jjy폴더에 "~월~일~년 점검결과.xlsx"로 저장

wb.save(new_filename)
# 여기까지 엑셀파일 생성 후 저장
#########################################################

wb = load_workbook(new_filename)  # new_file(위에서 만든 엑셀) 불러오기
# 엑셀파일 불러오는 경우 new_filename대신 엑셀파일 이름 넣기
# wb.save(new_filename)의 경우도 마찬가지
ws = wb.active

###시트지정
sh_list = wb.sheetnames  # 시트들을 리스트로 구분
w1 = wb[sh_list[0]]  # 첫번째 시트(report)를 w1로 지정
w2 = wb[sh_list[1]]  # 두번째 시트(log)를 w2로 지정
w3 = wb[sh_list[2]]  # 세번째 시트(action)를 w3으로 지정
w4 = wb[sh_list[3]]  # 네번째 시트(1. 계정관리)를 w4로 지정
w5 = wb[sh_list[4]]  # 다섯번째 시트(2. 파일 및 디렉토리 관리)를 w5로 지정
w6 = wb[sh_list[5]]  # 여섯번째 시트(3. 서비스 관리)를 w5로 지정
w7 = wb[sh_list[6]]  # 일곱번째 시트(4. 패치 관리)를 w6로 지정
w8 = wb[sh_list[7]]  # 여덟번째 시트(5. 로그 관리)를 w7로 지정

###계산
# ws["A6"] = "=SUM(A4:A5)"

left = Border(left=Side(style="thin"))
right = Border(right=Side(style="thin"))
top = Border(top=Side(style="thin"))
bottom = Border(bottom=Side(style="thin"))

##report 시트 항목 별 테두리
for i in range(4, 21, 4):
    for j in range(2, 7, 1):
        w1.cell(row=i, column=j).border = bottom

for i in range(3, 21, 4):
    for j in range(2, 7, 1):
        w1.cell(row=i, column=j).border = top
##가로줄 굵게
for j in range(3, 20, 4):
    for i in range(j, j + 2, 1):
        w1.cell(row=i, column=2).border = left

for j in range(3, 20, 4):
    for i in range(j, j + 2, 1):
        w1.cell(row=i, column=6).border = right
##세로줄 굵게

###reprot 시트 항목 별 테두리
for i in range(3, 20, 4):
    w1.cell(row=i, column=2).border = Border(
        top=Side(style="thin"),
        left=Side(style="thin"))
###좌상단 모서리 왼쪽, 위 굵게

for i in range(4, 21, 4):
    w1.cell(row=i, column=2).border = Border(
        bottom=Side(style="thin"),
        left=Side(style="thin"))
###좌하단 모서리 왼쪽, 아래 굵게

for i in range(3, 20, 4):
    w1.cell(row=i, column=6).border = Border(
        top=Side(style="thin"),
        right=Side(style="thin"))
###우상단 모서리 오른쪽, 위 굵게

for i in range(4, 21, 4):
    w1.cell(row=i, column=6).border = Border(
        bottom=Side(style="thin"),
        right=Side(style="thin"))
###우하단 모서리 오른쪽, 아래 굵게


####report 시트 종합 칸 테두리
for i in range(3, 8, 1):
    w1.cell(row=23, column=i).border = Border(top=Side(style="medium"))
    w1.cell(row=24, column=i).border = Border(top=Side(style="medium"))
    w1.cell(row=28, column=i).border = Border(bottom=Side(style="medium"))
####가로줄

for i in range(25, 28, 1):
    w1.cell(row=i, column=2).border = Border(left=Side(style="medium"))
    w1.cell(row=i, column=8).border = Border(right=Side(style="medium"))
####세로줄

####모서리 부분들
w1.cell(row=23, column=2).border = Border(
    top=Side(style="medium"),
    left=Side(style="medium"))
w1.cell(row=23, column=8).border = Border(
    top=Side(style="medium"),
    right=Side(style="medium"))

w1.cell(row=24, column=2).border = Border(
    top=Side(style="medium"),
    left=Side(style="medium"))
w1.cell(row=24, column=8).border = Border(
    top=Side(style="medium"),
    right=Side(style="medium"))

w1.cell(row=28, column=2).border = Border(
    bottom=Side(style="medium"),
    left=Side(style="medium"))
w1.cell(row=28, column=8).border = Border(
    bottom=Side(style="medium"),
    right=Side(style="medium"))

font1 = Font(size=12)  # font1 => 사이즈12
font2 = Font(size=14)  # font2 => 사이즈14
font3 = Font(size=14, bold=True)  # font2 => 사이즈14, 두껍게

#####report 시트 내용 작성

w1['B2'].value = '1. 계정 관리'
w1['B2'].font = font2
w1['B6'].value = '2. 파일 및 디렉토리 관리'
w1['B6'].font = font2
w1['B10'].value = '3. 서비스 관리'
w1['B10'].font = font2
w1['B14'].value = '4. 패치 관리'
w1['B14'].font = font2
w1['B18'].value = '5. 로그 관리'
w1['B18'].font = font2
# 1.계정관리부터 5.로그관리까지의 제목 추가

for i in range(3, 20, 4):
    w1.cell(row=i, column=2).value = '총 개수'
    w1.cell(row=i, column=2).font = font1
for i in range(3, 20, 4):
    w1.cell(row=i, column=3).value = '양호'
    w1.cell(row=i, column=3).font = font1
for i in range(3, 20, 4):
    w1.cell(row=i, column=4).value = '취약'
    w1.cell(row=i, column=4).font = font1
for i in range(3, 20, 4):
    w1.cell(row=i, column=5).value = 'N/A'
    w1.cell(row=i, column=5).font = font1
for i in range(3, 20, 4):
    w1.cell(row=i, column=6).value = '점수'
    w1.cell(row=i, column=6).font = font1
# 1.계정관리부터 5.로그관리까지의 항목 추가

w1.merge_cells("B23:D23")  # B23~D23까지의 셀 병합
# ws.unmerge_cells("A1:C1") #셀 분리

w1['B22'].value = '종합'
w1['B22'].font = font3
w1['B23'].value = '구분'
w1['E23'].value = '양호'
w1['F23'].value = '취약'
w1['G23'].value = 'N/A'
w1['H23'].value = '점수'
w1['B24'].value = '1. 계정 관리'
w1['B25'].value = '2. 파일 및 디렉토리 관리'
w1['B26'].value = '3. 서비스 관리'
w1['B27'].value = '4. 패치 관리'
w1['B28'].value = '5. 로그 관리'

wb.save(new_filename)
###엑셀 저장


# 텍스트 파일 읽을때 사용할 변수 선언
g = 1
b = 1

# instpect.txt 파일 한줄씩 읽어서 한줄씩 엑셀에 저장
try:
    file = open("C:\jjy\inspect.txt", 'r', encoding='utf-8')
    while True:
        line = file.readline()  # 파일을 읽어드려서 line으로 저장
        if line.startswith('A'):
            k = 1
            while w4.cell(row=k, column=1).value != None:
                k = k + 1
            w4.cell(row=k, column=1).value = line
            continue
        # if 읽어들인 line이 A로 시작하게 된다면 k에 1값을 넣어주고,
        # w1[A'k']셀이 빈칸이 아닐때 까지 k에 1씩 더해주고 읽어들이 값을 넣는다.
        elif line.startswith('B'):
            k = 1
            while w5.cell(row=k, column=1).value != None:
                k = k + 1
            w5.cell(row=k, column=1).value = line
            continue
        elif line.startswith('C'):
            k = 1
            while w6.cell(row=k, column=1).value != None:
                k = k + 1
            w6.cell(row=k, column=1).value = line
            continue
        elif line.startswith('D'):
            k = 1
            while w7.cell(row=k, column=1).value != None:
                k = k + 1
            w7.cell(row=k, column=1).value = line
            continue
        elif line.startswith('E'):
            k = 1
            while w8.cell(row=k, column=1).value != None:
                k = k + 1
            w8.cell(row=k, column=1).value = line
            continue
        else:
            break
    file.close()
    print("inspect.txt 저장 완료..")
except:
    try:
        file = open("C:\jjy\inspect.txt", 'r', encoding='ANSI')
        while True:
            line = file.readline()  # 파일을 읽어드려서 line으로 저장
            if line.startswith('A'):
                k = 1
                while w4.cell(row=k, column=1).value != None:
                    k = k + 1
                w4.cell(row=k, column=1).value = line
                continue
            # if 읽어들인 line이 A로 시작하게 된다면 k에 1값을 넣어주고,
            # w1[A'k']셀이 빈칸이 아닐때 까지 k에 1씩 더해주고 읽어들이 값을 넣는다.
            elif line.startswith('B'):
                k = 1
                while w5.cell(row=k, column=1).value != None:
                    k = k + 1
                w5.cell(row=k, column=1).value = line
                continue
            elif line.startswith('C'):
                k = 1
                while w6.cell(row=k, column=1).value != None:
                    k = k + 1
                w6.cell(row=k, column=1).value = line
                continue
            elif line.startswith('D'):
                k = 1
                while w7.cell(row=k, column=1).value != None:
                    k = k + 1
                w7.cell(row=k, column=1).value = line
                continue
            elif line.startswith('E'):
                k = 1
                while w8.cell(row=k, column=1).value != None:
                    k = k + 1
                w8.cell(row=k, column=1).value = line
                continue
            else:
                break
        file.close()
        print("inspect.txt 저장 완료..")
    except:
        print("inspect.txt 저장 실패")

# log.txt 파일 한줄씩 읽어서 한줄씩 엑셀에 저장
try:
    file = open("C:\jjy\log.txt", 'r', encoding='utf-8')
    while True:
        line = file.readline()
        if not line:
            break
        w2.cell(row=g, column=1).value = line  # line을 엑셀 두번째 시트에 데이터를 삽입
        g = g + 1
    file.close()
    print("log.txt 저장 성공..")
except:
    try:
        file = open("C:\jjy\log.txt", 'r', encoding='ANSI')
        while True:
            line = file.readline()
            if not line:
                break
            w2.cell(row=g, column=1).value = line  # line을 엑셀 두번째 시트에 데이터를 삽입
            g = g + 1
        file.close()
        print("log.txt 저장 성공..")
    except:
        print("log.txt 저장 실패..")

# action.txt파일 한줄씩 읽어서 한줄씩 엑셀에 저장
try:
    file = open("C:\jjy\Action.txt", 'r', encoding='utf-8')
    while True:
        line = file.readline()
        if not line:
            break
        w3.cell(row=b, column=1).value = line  # line을 엑셀 세번째 시트에 데이터를 삽입
        w1.cell(row=b + 49, column=1).value = line  # line을 엑셀 첫번째 시트, 그래프 아래에 데이터를 삽입
        b = b + 1
    file.close()
    print("action.txt 저장 성공..")
except:
    try:
        file = open("C:\jjy\Action.txt", 'r', encoding='ANSI')
        while True:
            line = file.readline()
            if not line:
                break
            w3.cell(row=b, column=1).value = line  # line을 엑셀 세번째 시트에 데이터를 삽입
            w1.cell(row=b + 49, column=1).value = line  # line을 엑셀 첫번째 시트, 그래프 아래에 데이터를 삽입
            b = b + 1
        file.close()
        print("action.txt 저장 성공..")
    except:
        print("action.txt 저장 실패..")

## inspect 시트 셀 나누기 ##
# A열에 있는 데이터를 column_a에 저장
column_a = w4['A']
column_b = w5['A']
column_c = w6['A']
column_d = w7['A']
column_e = w8['A']
# 저장된 데이터를 |를 기준으로 나누어서  values[0]을 기존 값에,
# values[1]을 2열에 출력
for cell in column_a:
    if cell.value:
        values = cell.value.split('|')
        cell.value = values[0].strip()
        w4.cell(row=cell.row, column=2, value=values[1].strip())

for cell in column_b:
    if cell.value:
        values = cell.value.split('|')
        cell.value = values[0].strip()
        w5.cell(row=cell.row, column=2, value=values[1].strip())

for cell in column_c:
    if cell.value:
        values = cell.value.split('|')
        cell.value = values[0].strip()
        w6.cell(row=cell.row, column=2, value=values[1].strip())

for cell in column_d:
    if cell.value:
        values = cell.value.split('|')
        cell.value = values[0].strip()
        w7.cell(row=cell.row, column=2, value=values[1].strip())

for cell in column_e:
    if cell.value:
        values = cell.value.split('|')
        cell.value = values[0].strip()
        w8.cell(row=cell.row, column=2, value=values[1].strip())

# 1.계정관리 셀 병합 및 수정
v0 = 1
while w4.cell(row=v0, column=1).value != None:
    v0 = v0 + 1
    w4.merge_cells(start_row=1, start_column=1, end_row= v0 - 1, end_column=1)
    w4['A1'].value = "1. 계정관리"
    w4['A1'].fill = PatternFill(fill_type='solid', fgColor=Color('789ABC'))

# 2. 파일 및 디렉토리 관리 셀 병합 및 셀 수정
v1 = 1
while w5.cell(row=v1, column=1).value != None:
    v1 = v1 + 1
    w5.merge_cells(start_row=1, start_column=1, end_row=v1 - 1, end_column=1)
    w5['A1'].value = "2. 파일 및 디렉토리 관리"
    w5['A1'].fill = PatternFill(fill_type='solid', fgColor=Color('789ABC'))

# 3. 서비스 관리 셀 병합 및 수정
v2 = 1
while w6.cell(row=v2, column=1).value != None:
    v2 = v2 + 1
    w6.merge_cells(start_row=1, start_column=1, end_row=v2 - 1, end_column=1)
    w6['A1'].value = "3. 서비스 관리"
    w6['A1'].fill = PatternFill(fill_type='solid', fgColor=Color('789ABC'))

# 4. 패치 관리 셀 병합 및 수정
v3 = 1
while w7.cell(row=v3, column=1).value != None:
    v3 = v3 + 1
    w7.merge_cells(start_row=1, start_column=1, end_row=v3 - 1, end_column=1)
    w7['A1'].value = "4. 패치 관리"
    w7['A1'].fill = PatternFill(fill_type='solid', fgColor=Color('789ABC'))

# 5. 로그 관리 셀 병합 및 수정
v4 = 1
while w8.cell(row=v4, column=1).value != None:
    v4 = v4 + 1
    w8.merge_cells(start_row=1, start_column=1, end_row=v4 - 1, end_column=1)
    w8['A1'].value = "5. 로그 관리"
    w8['A1'].fill = PatternFill(fill_type='solid', fgColor=Color('789ABC'))

# A열의 텍스트 중앙으로 정렬 및 폰트 변경
w4['A1'].alignment = Alignment(
    horizontal='center',
    vertical='center',
    wrap_text=True)
w4['A1'].font = Font(
    size=12, bold=True)
# A열의 셀 너비 13.1
w4.column_dimensions['A'].width = 13.1

w5['A1'].alignment = Alignment(
    horizontal='center',
    vertical='center',
    wrap_text=True)
w5['A1'].font = Font(
    size=12, bold=True)
w5.column_dimensions['A'].width = 13.1

w6['A1'].alignment = Alignment(
    horizontal='center',
    vertical='center',
    wrap_text=True)
w6['A1'].font = Font(
    size=12, bold=True)
w6.column_dimensions['A'].width = 13.1

w7['A1'].alignment = Alignment(
    horizontal='center',
    vertical='center',
    wrap_text=True)
w7['A1'].font = Font(
    size=12, bold=True)
w7.column_dimensions['A'].width = 13.1

w8['A1'].alignment = Alignment(
    horizontal='center',
    vertical='center',
    wrap_text=True)
w8['A1'].font = Font(
    size=12, bold=True)
w8.column_dimensions['A'].width = 13.1
###############################################################
###일단 여기까지 수정 완료


# report 시트 작성

# instpect 시트에서의 양호 및 취약 개수 세기
# 1. 계정 관리
w1['B4'].value = '=COUNTIF(inspect_1!B1:inspect_1!B100,"**")'
w1['C4'].value = '=COUNTIF(inspect_1!B1:inspect_1!B100,"*양호*")'
w1['D4'].value = '=COUNTIF(inspect_1!B1:inspect_1!B100,"*취약*")'
w1['E4'].value = '=COUNTIF(inspect_1!B1:inspect_1!B100,"*점검*")'
w1['F4'].value = '=IF(B4<1,0,ROUND(C4/B4*100,))'
# 2. 파일 및 디렉토리 관리
w1['B8'].value = '=COUNTIF(inspect_2!B1:inspect_2!B100,"**")'
w1['C8'].value = '=COUNTIF(inspect_2!B1:inspect_2!B100,"*양호*")'
w1['D8'].value = '=COUNTIF(inspect_2!B1:inspect_2!B100,"*취약*")'
w1['E8'].value = '=COUNTIF(inspect_2!B1:inspect_2!B100,"*점검*")'
w1['F8'].value = '=IF(B8<1,0,ROUND(C8/B8*100,))'
# 3. 서비스 관리
w1['B12'].value = '=COUNTIF(inspect_3!B1:inspect_3!B100,"**")'
w1['C12'].value = '=COUNTIF(inspect_3!B1:inspect_3!B100,"*양호*")'
w1['D12'].value = '=COUNTIF(inspect_3!B1:inspect_3!B100,"*취약*")'
w1['E12'].value = '=COUNTIF(inspect_3!B1:inspect_3!B100,"*점검*")'
w1['F12'].value = '=IF(B12<1,0,ROUND(C12/B12*100,))'
# 4. 패치 관리
w1['B16'].value = '=COUNTIF(inspect_4!B1:inspect_4!B100,"**")'
w1['C16'].value = '=COUNTIF(inspect_4!B1:inspect_4!B100,"*양호*")'
w1['D16'].value = '=COUNTIF(inspect_4!B1:inspect_4!B100,"*취약*")'
w1['E16'].value = '=COUNTIF(inspect_4!B1:inspect_4!B100,"*점검*")'
w1['F16'].value = '=IF(B16<1,0,ROUND(C16/B16*100,))'
# 5. 로그 관리
w1['B20'].value = '=COUNTIF(inspect_5!B1:inspect_5!B100,"**")'
w1['C20'].value = '=COUNTIF(inspect_5!B1:inspect_5!B100,"*양호*")'
w1['D20'].value = '=COUNTIF(inspect_5!B1:inspect_5!B100,"*취약*")'
w1['E20'].value = '=COUNTIF(inspect_5!B1:inspect_5!B100,"*점검*")'
w1['F20'].value = '=IF(B20<1,0,ROUND(C20/B20*100,))'
###########################
##########여기도 수정완료

# 종합
##1. 계정관리
w1['E24'].value = '=C4'
w1['F24'].value = '=D4'
w1['G24'].value = '=E4'
w1['H24'].value = '=F4'
##2. 파일 및 디렉토리 관리
w1['E25'].value = '=C8'
w1['F25'].value = '=D8'
w1['G25'].value = '=E8'
w1['H25'].value = '=F8'
# 3. 서비스 관리
w1['E26'].value = '=C12'
w1['F26'].value = '=D12'
w1['G26'].value = '=E12'
w1['H26'].value = '=F12'
# 4. 패치 관리
w1['E27'].value = '=C16'
w1['F27'].value = '=D16'
w1['G27'].value = '=E16'
w1['H27'].value = '=F16'
# 5. 로그 관리
w1['E28'].value = '=C20'
w1['F28'].value = '=D20'
w1['G28'].value = '=E20'
w1['H28'].value = '=F20'
# 총합
w1['B29'].value = '총합'
w1['D29'].value = '=SUM(B4, B8, B12, B16, B20)'
w1['E29'].value = '=SUM(E24:E28)'
w1['F29'].value = '=SUM(F24:F28)'
w1['G29'].value = '=SUM(G24:G28)'
w1['H29'].value = '=ROUND(E29/D29*100,0)'

# 총합 행 글자크기 12, 굵게
for i in range(2, 9, 1):
    w4.cell(row=29, column=i).font = Font(
        size=12, bold=True)

wb.save(new_filename)

w1['B32'].value = "구분"
w1['C32'].value = "총 개수"
w1['D32'].value = "양호 개수"
# 1.계정
w1['B33'].value = "=B24"
w1['C33'].value = "=B4"
w1['D33'].value = "=C4"
# 2.파일 및 디렉토리
w1['B34'].value = "=B25"
w1['C34'].value = "=B8"
w1['D34'].value = "=C8"
# 3.서비스
w1['B35'].value = "=B26"
w1['C35'].value = "=B12"
w1['D35'].value = "=C12"
# 4.패치
w1['B36'].value = "=B27"
w1['C36'].value = "=B16"
w1['D36'].value = "=C16"
# 5.로그
w1['B37'].value = "=B28"
w1['C37'].value = "=B20"
w1['D37'].value = "=C20"

w1['I6'].value = tday_s

chart = BarChart()
chart.type = "col"
chart.style = 40
chartData = Reference(w1, min_col=3, max_col=4, min_row=32, max_row=37)
category = Reference(w1, min_col=2, min_row=33, max_row=37)

chart.title = "취약점 진단 결과"
chart.add_data(chartData, titles_from_data=True)
chart.set_categories(category)

w1.add_chart(chart, 'B32')

##완료
#####################################################


wb.save(new_filename)
# os.startfile(new_filename) #엑셀 실행